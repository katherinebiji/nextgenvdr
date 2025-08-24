"""
Question processing service for parsing uploaded files and text input
"""
import openpyxl
import pypdf
from docx import Document as DocxDocument
import re
import uuid
from typing import List, Dict, Any, Tuple
from sqlalchemy.orm import Session
from models.schemas import QuestionCreate
from crud import create_question
import logging

logger = logging.getLogger(__name__)

class QuestionProcessor:
    def __init__(self):
        self.question_patterns = [
            r'^(\d+[\.\)])\s*(.+)$',  # 1. Question or 1) Question
            r'^[•\-\*]\s*(.+)$',      # • Question or - Question or * Question
            r'^([A-Z]\d*[\.\)])\s*(.+)$',  # A. Question or A1. Question
            r'^Question\s*\d*:?\s*(.+)$',  # Question: or Question 1:
        ]
    
    def extract_questions_from_text(self, text: str) -> List[str]:
        """Extract individual questions from text input"""
        lines = text.strip().split('\n')
        questions = []
        current_question = ""
        
        for line in lines:
            line = line.strip()
            if not line:
                if current_question.strip():
                    questions.append(current_question.strip())
                    current_question = ""
                continue
            
            # Check if line starts with a question pattern
            is_new_question = False
            for pattern in self.question_patterns:
                match = re.match(pattern, line, re.IGNORECASE)
                if match:
                    if current_question.strip():
                        questions.append(current_question.strip())
                    # Get the last capturing group (the question text)
                    current_question = match.groups()[-1] if match.groups() else line
                    is_new_question = True
                    break
            
            if not is_new_question:
                # Continue previous question or start new one
                if current_question:
                    current_question += " " + line
                else:
                    current_question = line
        
        # Add the last question
        if current_question.strip():
            questions.append(current_question.strip())
        
        # Filter out very short or invalid questions
        valid_questions = []
        for q in questions:
            if len(q) > 10 and ('?' in q or any(word in q.lower() for word in ['what', 'how', 'when', 'where', 'why', 'which', 'who', 'can', 'should', 'will', 'would', 'could', 'is', 'are', 'do', 'does', 'did'])):
                valid_questions.append(q)
        
        return valid_questions
    
    def extract_questions_from_excel(self, file_content: bytes) -> List[str]:
        """Extract questions from Excel file"""
        try:
            import io
            from openpyxl import load_workbook
            
            workbook = load_workbook(io.BytesIO(file_content))
            questions = []
            
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell and isinstance(cell, str):
                            # Try to extract questions from this cell
                            cell_questions = self.extract_questions_from_text(cell)
                            questions.extend(cell_questions)
            
            return questions
        except Exception as e:
            logger.error(f"Error processing Excel file: {e}")
            return []
    
    def extract_questions_from_word(self, file_content: bytes) -> List[str]:
        """Extract questions from Word document"""
        try:
            import io
            from docx import Document as DocxDocument
            
            doc = DocxDocument(io.BytesIO(file_content))
            text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            return self.extract_questions_from_text(text)
        except Exception as e:
            logger.error(f"Error processing Word file: {e}")
            return []
    
    def extract_questions_from_pdf(self, file_content: bytes) -> List[str]:
        """Extract questions from PDF file"""
        try:
            import io
            from pypdf import PdfReader
            
            reader = PdfReader(io.BytesIO(file_content))
            text = ""
            for page in reader.pages:
                text += page.extract_text() + "\n"
            
            return self.extract_questions_from_text(text)
        except Exception as e:
            logger.error(f"Error processing PDF file: {e}")
            return []
    
    def categorize_question(self, question: str) -> List[str]:
        """Categorize question based on content keywords"""
        question_lower = question.lower()
        categories = []
        
        # Financial categories
        if any(word in question_lower for word in ['revenue', 'profit', 'financial', 'ebitda', 'cash', 'debt', 'valuation', 'earnings', 'income']):
            categories.append('financial')
        
        # Legal categories  
        if any(word in question_lower for word in ['contract', 'legal', 'compliance', 'regulation', 'lawsuit', 'liability', 'intellectual property', 'patent', 'trademark']):
            categories.append('legal')
        
        # Operational categories
        if any(word in question_lower for word in ['operations', 'process', 'workflow', 'employee', 'staff', 'management', 'organization', 'customer', 'supplier']):
            categories.append('operational')
        
        # Technical categories
        if any(word in question_lower for word in ['technology', 'software', 'system', 'platform', 'infrastructure', 'data', 'security', 'api']):
            categories.append('technical')
        
        # Risk categories
        if any(word in question_lower for word in ['risk', 'threat', 'vulnerability', 'insurance', 'contingency', 'mitigation']):
            categories.append('risk')
        
        # Default category if none found
        if not categories:
            categories.append('general')
        
        return categories
    
    def process_file_upload(self, file_name: str, file_content: bytes, file_type: str) -> List[str]:
        """Process uploaded file and extract questions"""
        questions = []
        
        if file_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel']:
            questions = self.extract_questions_from_excel(file_content)
        elif file_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document', 'application/msword']:
            questions = self.extract_questions_from_word(file_content)
        elif file_type == 'application/pdf':
            questions = self.extract_questions_from_pdf(file_content)
        elif file_type == 'text/plain':
            text = file_content.decode('utf-8')
            questions = self.extract_questions_from_text(text)
        
        return questions
    
    def create_questions_from_upload(self, db: Session, questions: List[str], user_id: str, source_file: str = None) -> List[Dict[str, Any]]:
        """Create question records in database from extracted questions"""
        created_questions = []
        
        for i, question_text in enumerate(questions):
            # Generate title from first part of question
            title = question_text[:50] + "..." if len(question_text) > 50 else question_text
            
            # Categorize the question
            categories = self.categorize_question(question_text)
            
            question_data = QuestionCreate(
                title=title,
                content=question_text,
                priority="medium",  # Default priority
                tags=categories + ([f"source:{source_file}"] if source_file else [])
            )
            
            try:
                db_question = create_question(db, question_data, user_id)
                created_questions.append({
                    "id": db_question.id,
                    "title": db_question.title,
                    "content": db_question.content,
                    "status": db_question.status,
                    "priority": db_question.priority,
                    "tags": db_question.tags,
                    "asked_at": db_question.asked_at.isoformat()
                })
            except Exception as e:
                logger.error(f"Error creating question {i+1}: {e}")
                continue
        
        return created_questions

# Global instance
question_processor = QuestionProcessor()